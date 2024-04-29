import openpyxl
import os

# Load the Excel workbook
workbook = openpyxl.load_workbook('C:/Users/hecki/Downloads/layer.xlsx')

# Select the active sheet
sheet = workbook.active

# Get the values from the first row
first_row = sheet[1]

# Iterate over each cell in the first row
for cell in first_row:
    folder_name = str(cell.value).strip()
    
    # Replace invalid characters and spaces in the folder name
    folder_name = folder_name.replace(':', '_').replace(' ', '_')
    
    # Create the folder if it doesn't exist
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        print(f"Created folder: {folder_name}")
    else:
        print(f"Folder already exists: {folder_name}")
    
    # Get the column index of the current cell
    column_index = cell.column

    # Iterate over the cells in the column below the current cell
    for row in range(2, sheet.max_row + 1):
        subfolder_name = str(sheet.cell(row=row, column=column_index).value).strip()
        
        # Skip empty cells
        if not subfolder_name:
            continue
        
        # Replace invalid characters and spaces in the subfolder name
        subfolder_name = subfolder_name.replace(':', '_').replace(' ', '_')
        
        # Create the subfolder if it doesn't exist
        subfolder_path = os.path.join(folder_name, subfolder_name)
        if not os.path.exists(subfolder_path):
            os.makedirs(subfolder_path)
            print(f"Created subfolder: {subfolder_path}")
        else:
            print(f"Subfolder already exists: {subfolder_path}")
            #Test push after change